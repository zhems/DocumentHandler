from openpyxl import Workbook
from openpyxl import load_workbook 
import pandas as pd
import os
#"Monthly Report_Carpark_F1 - July 23.xlsx"

def format_excel(input_path,output_path):
    workbook = load_workbook(filename=input_path)
    sheet = workbook['Report A-Daily Counter Trans']
    floor = sheet.max_row

    Date = pd.Series([i[0].value for i in sheet[f'B10:B{floor}']],name='Date')
    Time = pd.Series([i[0].value for i in sheet[f'F10:F{floor}']],name='Time')
    Hourly_Entry = pd.Series([i[0].value for i in sheet[f'K10:K{floor}']],name='Hourly_Entry')
    Authorized_Entry = pd.Series([i[0].value for i in sheet[f'M10:M{floor}']],name='Authorized_Entry')
    Season_Entry = pd.Series([i[0].value for i in sheet[f'P10:P{floor}']],name='Season_Entry')
    Total_Entry = pd.Series([i[0].value for i in sheet[f'R10:R{floor}']],name='Total_Entry')

    Hourly_Exit = pd.Series([i[0].value for i in sheet[f'T10:T{floor}']],name='Hourly_Exit')
    CompPass_Exit = pd.Series([i[0].value for i in sheet[f'X10:X{floor}']],name='CompPass_Exit')
    Season_Exit = pd.Series([i[0].value for i in sheet[f'AA10:AA{floor}']],name='Season_Exit')
    Authorized_Exit = pd.Series([i[0].value for i in sheet[f'AE10:AE{floor}']],name='Authorized_Exit')
    Total_Exit = pd.Series([i[0].value for i in sheet[f'AG10:AG{floor}']],name='Total_Exit')

    Season_Util_Rate = pd.Series([i[0].value for i in sheet[f'AJ10:AJ{floor}']],name='Season_Util_Rate')
    Shortterm_Util_Rate = pd.Series([i[0].value for i in sheet[f'AL10:AL{floor}']],name='Shortterm_Util_Rate')
    Hourly_Util_Rate = pd.Series([i[0].value for i in sheet[f'AN10:AN{floor}']],name='Hourly_Util_Rate')

    df = pd.concat([Hourly_Entry,Authorized_Entry,Season_Entry,Total_Entry,
            Hourly_Exit,CompPass_Exit,Season_Exit,Authorized_Exit,Total_Exit,
            Season_Util_Rate,Shortterm_Util_Rate,Hourly_Util_Rate], axis=1)
    
    df.to_csv(output_path,index=False)

if __name__ == '__main__':
    for f in os.listdir('input'):
        format_excel('input/'+f,f"output/{f.split('-')[1].split('.')[0].strip()}.csv")
        os.replace('input/'+f,'processed_input/'+f)